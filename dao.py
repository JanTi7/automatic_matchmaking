import functools
import sys
import time
import pathlib
import logging
from tinydb import TinyDB, Query
from dataclasses import dataclass, asdict, field
from typing import ClassVar, Optional, Callable
from datetime import timedelta
from functools import cache


from rich.console import Console
from rich.columns import Columns
from rich.panel import Panel

from matching_algos.base_matching_algo import BaseMatchingAlgo

from collections import namedtuple, defaultdict

from matching_algos.interactive_matching import InteractiveMatcher
from matching_algos.task_input_output import TaskOutput

Participants = namedtuple("Participants", "a1 a2 b1 b2")

db = None
db_name = None
id2name = dict()


def load_default_database(create_new=False):
    DB_DIR = pathlib.Path("databases")
    use_database((DB_DIR / ".default").read_text().strip(), create_new=create_new)


def set_default_database(db_name):
    DB_DIR = pathlib.Path("databases")
    (DB_DIR / ".default").write_text(db_name)


def use_database(
    db_name, create_new=False, init_logging=True, read_only=False, print_ascii=True
):
    if print_ascii:
        import pyfiglet

        pyfiglet.print_figlet(f"Using database: {db_name}")

    # print(f"Using database: {db_name}")
    DB_DIR = pathlib.Path("databases")
    FILE_TO_LOAD = DB_DIR / db_name

    if not create_new:
        if not FILE_TO_LOAD.exists():
            raise FileNotFoundError(
                f"There is no database at {FILE_TO_LOAD}. Try creating one first or use a flag."
            )

    FILE_TO_LOAD.parent.mkdir(parents=True, exist_ok=True)

    read_only_flag = dict(access_mode="r") if read_only else dict()

    globals()["db_name"] = db_name
    globals()["db"] = TinyDB(
        FILE_TO_LOAD,
        ensure_ascii=False,
        sort_keys=True,
        indent=4,
        separators=(",", ": "),
        **read_only_flag,
    )

    if init_logging:
        import datetime
        import socket

        filename = f'logs/{db_name.split(".")[0]}_{datetime.date.today()}_{socket.gethostname()}.log'
        pathlib.Path(filename).parent.mkdir(parents=True, exist_ok=True)
        logging.basicConfig(
            filename=filename,
            encoding="utf-8",
            level=logging.DEBUG,
            format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        )

        def handle_exception(exc_type, exc_value, exc_traceback):
            if issubclass(exc_type, KeyboardInterrupt):
                sys.__excepthook__(exc_type, exc_value, exc_traceback)
                return

            logging.critical(
                "Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback)
            )
            print((exc_type, exc_value, exc_traceback))

        sys.excepthook = handle_exception

        logging.info(f"Using database: {db_name}")

    clear_id2name_cache()


def clear_id2name_cache():
    logging.debug("Clearing caches")
    _generate_playerid_to_uniquename_map.cache_clear()


Entry = Query()


@dataclass
class Player:
    first_name: str
    last_name: str
    rating_snapshot: str
    games_played: int
    games_paused: int
    games_won: int
    games_lost: int
    total_rating_change: int

    TABLE_NAME: ClassVar[str] = "player"
    player_id: str = field(default_factory=lambda: generate_id(Player.TABLE_NAME))

    def get_name(self, depth=1, bold_first_name=False):
        fn = f"[b]{self.first_name}[/b]" if bold_first_name else self.first_name

        if depth == 0:
            return fn

        return f"{fn} {self.last_name[:depth]}."

    def get_rating_snapshot(self):
        rating_table = db.table(RatingSnapshot.TABLE_NAME)
        rating = rating_table.search(Entry.rating_snap_id == self.rating_snapshot)[0]
        return RatingSnapshot(**rating)

    def get_current_rating(self):
        return self.get_rating_snapshot().rating  # + numpy.random.normal(0, 0.5)

    def to_panel(self, id2name: dict, style="none"):
        return Panel(
            f"[b]{id2name[self.player_id]}[/b]",
            subtitle=str(self.get_current_rating()),
            style=style,
        )


def get_player_from_id(player_id):
    player_table = db.table(Player.TABLE_NAME)
    try:
        player_dict = player_table.search(Entry.player_id == player_id)[0]
    except IndexError:
        raise IndexError(
            f"No player with id {repr(player_id)}. Player ids are {[d['player_id'] for d in player_table.all()]}"
        )
    return Player(**player_dict)


def get_all_players():
    player_table = db.table(Player.TABLE_NAME)

    return [Player(**e) for e in player_table.all()]


def generate_playerid_to_uniquename_map(list_of_pids: list, bold_first_name=False):
    tuple_of_pids = tuple(sorted(list_of_pids))
    return _generate_playerid_to_uniquename_map(tuple_of_pids, bold_first_name)


@cache
def _generate_playerid_to_uniquename_map(tuple_of_pids: tuple, bold_first_name=False):
    name2id = defaultdict(set)

    # if the list is empty, assume all pids
    if len(tuple_of_pids) > 0:
        player_pool = [get_player_from_id(pid) for pid in tuple_of_pids]
    else:
        player_pool = get_all_players()

    for player in player_pool:
        name2id[player.get_name(depth=0, bold_first_name=bold_first_name)].add(
            player.player_id
        )

    names_to_remove = list()
    items_to_add = list()
    for name, pids in name2id.items():
        if len(pids) == 1:
            continue

        names_to_remove.append(name)
        final_depth = None
        for depth in range(1, 99):
            new_names = {
                get_player_from_id(pid).get_name(depth, bold_first_name=bold_first_name)
                for pid in pids
            }
            if len(new_names) == len(pids):
                final_depth = depth
                break

        for pid in pids:
            items_to_add.append(
                (
                    get_player_from_id(pid).get_name(
                        final_depth, bold_first_name=bold_first_name
                    ),
                    pid,
                )
            )

    for name_to_remove in names_to_remove:
        name2id.pop(name_to_remove)

    for name, pid in items_to_add:
        name2id[name].add(pid)

    id2name = {list(val)[0]: key for key, val in name2id.items()}

    globals()["id2name"].update(id2name)

    return id2name


def time_adjust_rd(
    old_rd, time_passed: timedelta, max_rd=155, inc_constant=25, max_inc=100
):
    if old_rd > max_rd:
        return old_rd

    weeks_passed = time_passed.days // 7
    if time_passed.days % 7 >= 5:
        weeks_passed += 1

    from math import sqrt

    rd = old_rd
    for _ in range(weeks_passed):
        rd = min(sqrt(rd**2 + inc_constant**2), max_rd)

    diff = rd - old_rd

    inc = min(max_inc, diff)  # after 4 weeks it doesn't matter if you stay away longer
    new_rd = min(max_rd, old_rd + inc)

    return new_rd


@dataclass
class StartOfPeriodRating:
    rating_snap_id: str

    previous_period_id: str

    TABLE_NAME: ClassVar[str] = "pe_rating"
    pe_rating_id: str = field(
        default_factory=lambda: generate_id(StartOfPeriodRating.TABLE_NAME)
    )


@dataclass
class IntermediateRating:
    rating_snap_id: str
    pe_rating_id: str

    TABLE_NAME: ClassVar[str] = "inter_rating"
    inter_rating_id: str = field(
        default_factory=lambda: generate_id(IntermediateRating.TABLE_NAME)
    )


@dataclass
class RatingSnapshot:
    player_id: str
    timestamp: float
    rating: int
    rd: float
    vol: float

    game_res_id: str  # the id of the game which yielded the new snapshot

    parent_id: str = "None"  # either a StartOfPeriodRating or an IntermediateRating

    TABLE_NAME: ClassVar[str] = "rating_snap"
    rating_snap_id: str = field(
        default_factory=lambda: generate_id(RatingSnapshot.TABLE_NAME)
    )

    def __post_init__(self):
        self.rating = int(round(self.rating, 0))

    def rd_time_adjusted(self):
        return time_adjust_rd(self.rd, timedelta(seconds=time.time() - self.timestamp))

    def to_glicko2_player(self):
        from algo_glicko2 import Glicko2Player

        return Glicko2Player(
            rating=self.rating, rd=self.rd, vol=self.vol, pid=self.player_id
        )


@dataclass
class BlockOfGames:
    proposed: dict  # of pp_ids
    players_pausing: list  # of player_ids
    # players_present   list  # of player_ids which paused|played, to easily find all blocks where a player was present
    results: dict = field(default_factory=dict)  # of res_ids
    cancelled: dict = field(default_factory=dict)  # of pp_ids

    committed: bool = False

    timestamp: float = field(default_factory=time.time)
    TABLE_NAME: ClassVar[str] = "block_of_games"
    block_of_games_id: str = field(
        default_factory=lambda: generate_id(BlockOfGames.TABLE_NAME)
    )

    def draw(self):
        from viz import viz_block_of_games

        viz_block_of_games(self)

    def register_result(
        self, local_game_idx, points_a, points_b, timestamp: Optional[float] = None
    ):
        game = GameProposed(**load_from_db(self.proposed[local_game_idx]))
        pp_dict = asdict(game)
        pp_dict.pop("game_pp_id")
        pp_dict.pop("timestamp")

        if timestamp:
            pp_dict["timestamp"] = timestamp

        game_res = GameResult(points_a=points_a, points_b=points_b, **pp_dict)
        self.results[local_game_idx] = game_res.game_res_id
        del self.proposed[local_game_idx]
        save_to_db(game_res)
        save_to_db(self, update_if_exists=True)

    def unregister_result(self, local_game_idx):
        game = GameResult(**load_from_db(self.results[local_game_idx]))
        pp_dict = asdict(game)
        pp_dict.pop("game_res_id")
        pp_dict.pop("timestamp")
        pp_dict.pop("points_a")
        pp_dict.pop("points_b")

        game_res = GameProposed(**pp_dict)
        self.proposed[local_game_idx] = game_res.game_pp_id
        del self.results[local_game_idx]
        save_to_db(game_res)
        save_to_db(self, update_if_exists=True)

    def commit(self):
        if len(self.results) > 0:
            # if all games were cancelled, nobody paused, so make sure at least one finished normally
            for p_id in self.players_pausing:
                player = get_player_from_id(p_id)
                player.games_paused += 1
                save_to_db(player, update_if_exists=True)

        for game_res_id in self.results.values():
            game_res = GameResult(**load_from_db(game_res_id))

            new_ratings = game_res.ratings_post_result()
            for new_rating in new_ratings:
                save_to_db(new_rating)
                player = get_player_from_id(new_rating.player_id)
                player_old_rating = player.get_current_rating()

                player.rating_snapshot = new_rating.rating_snap_id
                player.games_played += 1

                if game_res.did_player_win(new_rating.player_id):
                    player.games_won += 1
                else:
                    player.games_lost += 1
                player.total_rating_change += new_rating.rating - player_old_rating

                save_to_db(player, update_if_exists=True)

        self.committed = True
        save_to_db(self, update_if_exists=True)


def save_block_of_games(block: BlockOfGames):
    pool_table = db.table(BlockOfGames.TABLE_NAME)
    pool_table.insert(asdict(block))


def load_unfinished_blocks_of_games():
    block_table = db.table(BlockOfGames.TABLE_NAME)
    return [
        BlockOfGames(**entry) for entry in block_table.search(Entry.committed == False)
    ]


def remove_all_unfinished_blocks_of_games():
    block_table = db.table(BlockOfGames.TABLE_NAME)
    # entries_to_del = block_table.search(Entry.committed == False)
    block_table.remove(Entry.committed == False)


@dataclass
class GameProposed:
    rsnap_a_1: str
    rsnap_a_2: str
    rsnap_b_1: str
    rsnap_b_2: str

    importance: float = 1.0

    TABLE_NAME: ClassVar[str] = "game_pp"
    game_pp_id: str = field(
        default_factory=lambda: generate_id(GameProposed.TABLE_NAME)
    )
    timestamp: float = field(default_factory=time.time)

    def draw(self, local_idx=""):
        from viz import viz_single_game

        viz_single_game(self, local_idx=local_idx)

    def players(self):
        return Participants(
            *_rsnaps_to_pids(
                self.rsnap_a_1, self.rsnap_a_2, self.rsnap_b_1, self.rsnap_b_2
            )
        )


@functools.cache
def _rsnap_to_pid(rsnap_id):
    return load_from_db(rsnap_id)["player_id"]


@functools.cache
def _rsnaps_to_pids(*args):
    return [_rsnap_to_pid(rs) for rs in args]


@dataclass
class GameResult:
    rsnap_a_1: str
    rsnap_a_2: str
    rsnap_b_1: str
    rsnap_b_2: str

    points_a: int
    points_b: int

    importance: float = 1.0

    TABLE_NAME: ClassVar[str] = "game_res"
    game_res_id: str = field(default_factory=lambda: generate_id(GameResult.TABLE_NAME))
    timestamp: float = field(default_factory=time.time)

    def __post_init__(self):
        self.points_a = int(self.points_a)
        self.points_b = int(self.points_b)

    def draw(self, local_idx=""):
        from viz import viz_single_game

        viz_single_game(self, local_idx=local_idx)

    def did_player_win(self, player_id):
        if player_id in (self.players().a1, self.players().a2):
            return self.points_a > self.points_b
        elif player_id in (self.players().b1, self.players().b2):
            return self.points_a < self.points_b
        else:
            raise ValueError(
                f"Player with id {player_id} did not participate in game {asdict(self)}"
            )

    def ratings_pre_result(self):
        return tuple(
            RatingSnapshot(**load_from_db(rsnap_id))
            for rsnap_id in (
                self.rsnap_a_1,
                self.rsnap_a_2,
                self.rsnap_b_1,
                self.rsnap_b_2,
            )
        )

    def ratings_post_result(self, log=True):
        rating_a_1 = RatingSnapshot(**load_from_db(self.rsnap_a_1))
        player = get_player_from_id(rating_a_1.player_id)

        rating_a_2 = RatingSnapshot(**load_from_db(self.rsnap_a_2))
        player1 = get_player_from_id(rating_a_2.player_id)

        rating_b_1 = RatingSnapshot(**load_from_db(self.rsnap_b_1))
        player2 = get_player_from_id(rating_b_1.player_id)

        rating_b_2 = RatingSnapshot(**load_from_db(self.rsnap_b_2))
        player3 = get_player_from_id(rating_b_2.player_id)

        from algo_glicko2 import register_game_result

        if log:
            logging.info(
                f"Registering game result from {id2name.get(player.player_id, player.player_id)} & {id2name.get(player1.player_id, player1.player_id)} vs {id2name.get(player2.player_id, player2.player_id)} & {id2name.get(player3.player_id, player3.player_id)}"
            )
            logging.info(f"Match result was {self.points_a} - {self.points_b}")

        return register_game_result(
            self.game_res_id,
            rating_a_1,
            rating_a_2,
            rating_b_1,
            rating_b_2,
            points_a=self.points_a,
            points_b=self.points_b,
        )

    def players(self):
        return Participants(
            *_rsnaps_to_pids(
                self.rsnap_a_1, self.rsnap_a_2, self.rsnap_b_1, self.rsnap_b_2
            )
        )


def get_all_games_in_timewindow(delta: timedelta, verbose=False):
    gameblock_table = db.table(BlockOfGames.TABLE_NAME)
    all_relevant_blocks = gameblock_table.search(
        (Entry.timestamp >= time.time() - delta.total_seconds())
        & (Entry.committed == True)
    )

    games = list()
    for block in [BlockOfGames(**data) for data in all_relevant_blocks]:
        for game in block.results.values():
            game_res = GameResult(**load_from_db(game))
            games.append(game_res)

    return games


def get_participants_from_last_matches(delta: timedelta, verbose=False):
    all_tuples = list()

    gameblock_table = db.table(BlockOfGames.TABLE_NAME)
    all_relevant_blocks = gameblock_table.search(
        (Entry.timestamp >= time.time() - delta.total_seconds())
        & (Entry.committed == True)
    )

    if verbose:
        print("[DAO] [played together] Blocks:", all_relevant_blocks)

    # gameblocks from last 24 hours
    for block in [BlockOfGames(**data) for data in all_relevant_blocks]:
        block_list = list()
        for game in block.results.values():
            game_res = GameResult(**load_from_db(game))
            block_list.append(game_res.players())

        if len(block_list) > 0:
            all_tuples.append((block.timestamp, block_list))

    if verbose:
        print("[DAO] [played together last 24h]", all_tuples)
    return all_tuples


def _number_of_sets(num_sets, num_players):
    from prompt_toolkit import prompt

    if num_sets is None:
        N_SETS = int(num_players // 4)
    elif num_sets == "interactive":
        N_SETS = int(num_players // 4)
        N_SETS = int(prompt("Number of sets: ", default=str(N_SETS)))
    elif type(num_sets) == int:
        N_SETS = num_sets
    else:
        N_SETS = int(num_sets)

    return N_SETS


def get_matching_algo_from_str(matching_algo_str) -> Callable[[], BaseMatchingAlgo]:
    from matching_algos.bruteforce_matching_minizinc import BruteforceMatcherMinizinc
    from matching_algos.bruteforce_matching import BruteforceMatcher
    from matching_algos.random_matching import RandomMatcher

    algo_dict = dict(
        default=BruteforceMatcher,  # set as default since it doesn't require an external program - minizinc runs more reliably
        random=lambda: RandomMatcher(tries=3),
        interactive=InteractiveMatcher,
        scipy=BruteforceMatcher,
        minizinc=BruteforceMatcherMinizinc,
    )
    algo_dict.update(
        studientermin1=algo_dict["random"],
        studientermin2=algo_dict["default"],
        studientermin3=algo_dict["random"],
        studientermin4=algo_dict["default"],
    )
    logging.debug(f"Using matching algo {repr(matching_algo_str)}.")
    print(f"Using matching algo {repr(matching_algo_str)}.")
    try:
        return algo_dict[matching_algo_str]
    except KeyError:
        logging.error(
            f"Could not find matching algo with name {repr(matching_algo_str)}. "
            f"Using default algo. Available Options are {list(algo_dict.keys())}."
        )
        return algo_dict["default"]


@dataclass
class PlayerPool:
    timestamp: float = field(default_factory=time.time)
    list_of_player_ids: list = field(default_factory=list)
    players_voluntarily_pausing: list = field(default_factory=list)

    TABLE_NAME: ClassVar[str] = "player_pool"
    player_pool_id: str = field(
        default_factory=lambda: generate_id(PlayerPool.TABLE_NAME)
    )

    def __len__(self):
        """Returns the number of players willing to play the next round."""
        return len(self.list_of_player_ids)

    def _save_changes(self):
        self.timestamp = time.time()
        save_to_db(self, update_if_exists=True)

    def everybody(self) -> list[str]:
        return self.list_of_player_ids + self.players_voluntarily_pausing

    def add_player(self, player_id: str):
        self.list_of_player_ids.append(player_id)
        self._save_changes()

    def remove_player(self, player_id: str):
        if player_id in self.list_of_player_ids:
            self.list_of_player_ids.remove(player_id)
        elif player_id in self.players_voluntarily_pausing:
            self.players_voluntarily_pausing.remove(player_id)
        else:
            print(f"Can't remove player {player_id} - not in pool!")
        self._save_changes()

    def remove_all(self):
        self.list_of_player_ids.clear()
        self.players_voluntarily_pausing.clear()
        self._save_changes()

    def pause_player(self, player_id: str):
        self.list_of_player_ids.remove(player_id)
        self.players_voluntarily_pausing.append(player_id)
        self._save_changes()

    def unpause_player(self, player_id: str):
        self.players_voluntarily_pausing.remove(player_id)
        self.list_of_player_ids.append(player_id)
        self._save_changes()

    ################################################################

    def draw(self):
        id2name = generate_playerid_to_uniquename_map(self.everybody())

        console = Console(width=min(Console().width, 100))
        cwidth = console.width
        unpaused_panels = [
            get_player_from_id(p_id).to_panel(id2name)
            for p_id in self.list_of_player_ids
        ]
        paused_panels = [
            get_player_from_id(p_id).to_panel(id2name, style="on #aa2222")
            for p_id in self.players_voluntarily_pausing
        ]
        console.print(
            Columns(unpaused_panels + paused_panels, width=(cwidth - 16) // 4)
        )

    def preview_pause(self, pause_mode, num_sets=None):
        from prompt_toolkit import prompt
        from find_matchups import select_players_to_pause

        players = [get_player_from_id(pid) for pid in self.list_of_player_ids]
        N_PLAYERS = len(players)
        if num_sets is None:
            N_SETS = int(N_PLAYERS // 4)
            N_SETS = int(prompt("Number of sets: ", default=str(N_SETS)))
        else:
            N_SETS = num_sets
        select_players_to_pause(players, N_PLAYERS - N_SETS * 4, pause_mode)

    def start_next_round(
        self,
        pause_mode,
        higher_rating_weight,
        num_sets=None,
        matching_algo: Optional[str | Callable[[], BaseMatchingAlgo]] = None,
        log_task=False,
        return_task_output=False,
    ) -> BlockOfGames | tuple[BlockOfGames, TaskOutput]:
        from find_matchups import (
            select_players_to_pause,
            find_and_viz_solution,
            WeightMatrixManager,
        )

        players = [get_player_from_id(pid) for pid in self.list_of_player_ids]
        assert len(players) >= 4, "Not enough players to play a single game!"
        players.sort(key=lambda p: p.get_current_rating(), reverse=True)
        N_PLAYERS = len(players)
        N_SETS = _number_of_sets(num_sets, N_PLAYERS)

        if matching_algo is None:
            matcher = get_matching_algo_from_str("default")()
        else:
            if type(matching_algo) is str:
                matcher = get_matching_algo_from_str(matching_algo)()
            else:
                matcher = matching_algo()

        players_to_play, players_to_pause = select_players_to_pause(
            players, N_PLAYERS - N_SETS * 4, pause_mode
        )

        matrix_manager = WeightMatrixManager(players_to_play)
        task_input = matrix_manager.get_task_input(
            higher_rating_weight=higher_rating_weight
        )
        task_output = matcher.find_matching(task_input, log_task=log_task)
        # task_output.viz_result(pathlib.Path("/tmp/manager_viz"))

        proposed_games = task_output.convert_to_proposed_games(matrix_manager)
        for game in proposed_games:
            save_to_db(game)
        gameblock = BlockOfGames(
            proposed={str(idx): g.game_pp_id for idx, g in enumerate(proposed_games)},
            players_pausing=[p.player_id for p in players_to_pause]
            + self.players_voluntarily_pausing,
        )
        save_to_db(gameblock)

        if return_task_output:
            return gameblock, task_output
        else:
            return gameblock


def save_to_db(data, update_if_exists=False, verbose=False):
    if verbose:
        from pprint import pprint

        pprint(data)

    table = db.table(data.TABLE_NAME)
    id_name = f"{data.TABLE_NAME}_id"

    if update_if_exists:
        assert hasattr(data, id_name)
        table.upsert(asdict(data), Entry[id_name] == getattr(data, id_name))
    else:
        if hasattr(data, id_name):
            assert (
                len(table.search(Entry[id_name] == getattr(data, id_name))) == 0
            ), "Entry with this id already exists"
        table.insert(asdict(data))


def load_from_db(id_str):
    # print("LOAD FROM DB", repr(id_str))
    # print(f"Trying to load {id_str} from db!")
    table_name = id_str[: id_str.rfind("_")]  # Should always be right in our case
    id_name = table_name + "_id"
    table = db.table(table_name)
    try:
        return table.search(Entry[id_name] == id_str)[0]
    except IndexError:
        raise IndexError(
            f"Table {table_name} does not contain an element with {id_name} {repr(id_str)}"
        )


def load_last_player_pool():
    pool_table = db.table(PlayerPool.TABLE_NAME)
    try:
        latest_pool = sorted(pool_table.all(), key=lambda pd: pd["timestamp"])[-1]
        return PlayerPool(**latest_pool)
    except IndexError:
        return None


def load_player_pool_interactive():
    last_player_pool = load_last_player_pool()
    if last_player_pool is None:
        print("No old player pools found!")
        return None

    time_passed = timedelta(seconds=time.time() - last_player_pool.timestamp)
    if time_passed <= timedelta(minutes=180):
        return last_player_pool

    from prompt_toolkit.shortcuts import confirm

    if confirm(f"The last pool is {time_passed} old. Do you wish to load it? "):
        return last_player_pool
    return None


def generate_id(prefix):
    from collections import defaultdict

    prefix2nbytes = defaultdict(lambda: 4)
    prefix2nbytes["player"] = 2

    table = db.table(prefix)
    set_of_other_ids = {e.get(f"{prefix}_id", None) for e in table.all()}

    import secrets

    while True:
        id = f"{prefix}_{secrets.token_hex(prefix2nbytes[prefix])}"
        if id not in set_of_other_ids:
            return id


def add_new_player(first_name, last_name, init_rating, init_rd, player_id=None):
    player_table = db.table(Player.TABLE_NAME)

    ## check if player with same name is already in system

    same_name = player_table.search(
        (Entry.first_name == first_name) & (Entry.last_name == last_name)
    )
    if len(same_name) > 0:
        print(
            "THERE ALREADY IS A PLAY WITH THE EXACT SAME NAMES IN THE SYSTEM, SKIPPING!"
        )
        return same_name[0]["player_id"]

    new_id = (
        f"{Player.TABLE_NAME}_{player_id}"
        if player_id is not None
        else generate_id(Player.TABLE_NAME)
    )
    # new_id = generate_id(Player.TABLE_NAME)  # needs to be done explicitly here bc we need it for the snapshot

    initial_rating_snapshot = RatingSnapshot(
        player_id=new_id,
        timestamp=time.time(),
        rating=init_rating,
        rd=init_rd,
        vol=0.06,
        game_res_id="__INIT__",
    )

    new_player = Player(
        player_id=new_id,
        first_name=first_name.strip(),
        last_name=last_name.strip(),
        games_played=0,
        games_paused=0,
        rating_snapshot=initial_rating_snapshot.rating_snap_id,
        games_lost=0,
        games_won=0,
        total_rating_change=0,
    )

    save_to_db(initial_rating_snapshot)
    save_to_db(new_player)

    clear_id2name_cache()

    return new_player.player_id


def get_all_names():
    player_table = db.table(Player.TABLE_NAME)

    all_names = [
        (e["first_name"] + " " + e["last_name"], e["player_id"])
        for e in player_table.search(Entry.player_id.exists())
    ]

    if len(all_names) == 0:
        return list(), list()
    names, ids = zip(*all_names)

    return list(names), list(ids)


from rating_algos.base_rating_algo import X_data, Y_data


def get_games_for_prediction() -> (list[list[X_data]], list[list[Y_data]]):
    x, y = list(), list()
    rating_estimates = list()
    block_table = db.table(BlockOfGames.TABLE_NAME)

    all_blocks = sorted(block_table.all(), key=lambda d: d.doc_id)

    for block in [
        block
        for b in all_blocks
        if (block := BlockOfGames(**b)) and len(block.results) > 0
    ]:
        x.append(
            [
                X_data(*g.players(), g.timestamp, g.game_res_id)
                for game in block.results.values()
                if (g := GameResult(**load_from_db(game)))
            ]
        )
        y.append(
            [
                Y_data(g.points_a, g.points_b)
                for game in block.results.values()
                if (g := GameResult(**load_from_db(game)))
            ]
        )

        rating_estimates.append(
            [
                (g.game_res_id,) + tuple(g.ratings_pre_result())
                for game in block.results.values()
                if (g := GameResult(**load_from_db(game)))
            ]
        )

    return x, y, rating_estimates


def get_initial_rating_estimates() -> dict[str, int]:
    q = Query()

    rs_table = db.table(RatingSnapshot.TABLE_NAME)

    all_initial_snaps = rs_table.search(q.game_res_id == "__INIT__")

    return {snap["player_id"]: int(snap["rating"]) for snap in all_initial_snaps}


def import_from_excel(filepath, start_rd=125):
    from openpyxl import load_workbook

    wb = load_workbook(filename=filepath)
    sheet = wb.active

    current_cell = sheet.cell(3, 1)

    players_added = 0

    while current_cell.value is not None:
        # print(current_cell.value)
        name = current_cell.offset(0, 1).value
        rating = current_cell.offset(0, 2).value

        if name is not None and rating is not None:
            try:
                firstname, lastname = name.rsplit(" ", maxsplit=1)
            except ValueError:
                firstname, lastname = name, "1"

            add_new_player(
                firstname, lastname, rating, start_rd, player_id=current_cell.value
            )
            players_added += 1

            print("Adding", firstname, lastname, int(rating))

        elif (name, rating) != (None, None):
            logging.warning(
                f"Zeile {current_cell.value} unvollständig. Skipping. {name=} {rating=}"
            )

        current_cell = current_cell.offset(1, 0)

    print(f"Added {players_added} players.")
