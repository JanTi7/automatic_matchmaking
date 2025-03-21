import pathlib
from argparse import ArgumentParser


def get_db_name(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    sheet = wb.active
    db_name = sheet.cell(1, 7).value
    return db_name


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("excel_file")
    parser.add_argument("--init_rd", type=int, default=125)


    args = parser.parse_args()

    db_name = get_db_name(args.excel_file)
    print(f"{db_name=}")

    import shutil
    from helper import get_timestamp

    db_path = pathlib.Path(f"databases/{db_name}")
    if db_path.exists():
        new_path = db_path.parent / f"{get_timestamp()}_{db_name}"
        print(f"Moving already existing db to {new_path}")
        shutil.move(db_path, new_path)

    from dao import use_database, import_from_excel, set_default_database, create_default_table_config

    use_database(db_name, create_new=True)
    import_from_excel(args.excel_file, args.init_rd)

    set_default_database(db_name)
    create_default_table_config(save_to_database=True)
